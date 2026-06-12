"""Industry data loader for CFM and TrendForce semiconductor data.

CFM: Each sheet = one product spec with standard header (Base Date, Release Date, Time, TZ, Low, High, Average).
     Loads the Average column per sheet.

TrendForce: Two-row header. Row 0 = product name (repeated per H/L/A group), Row 1 = High/Low/Average.
     Auto-discovers products and loads the Average column for each.

Special cases handled:
- CFM Index.xlsx: simple columns (DRAM Index, NAND Index), not Low/High/Average
- CFM Flash Wafer: uses Low/Open/Close instead of Low/High/Average (loads Close)
- TrendForce Index.xlsx (DXI): single-row header, standard format
- TrendForce Li-Ion Battery: single-row header with descriptive column names
- TrendForce LCD: has Low/High/Average/Last Avg (4 cols per product)
"""
import re
import sys
from pathlib import Path
from datetime import datetime, date

import duckdb
import openpyxl

DB_PATH = Path(__file__).parent / 'alternative_data.duckdb'
DATA_DIR = Path(__file__).parent / 'data' / 'industry'

SKIP_SHEETS = {'fig', 'info', 'Info'}


def _parse_date(raw) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return None


def _slugify(name: str) -> str:
    """Convert product name to UPPER_SNAKE indicator_id component."""
    s = name.strip().replace('\xa0', '')
    s = re.sub(r'[()/:,\-–—\s]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s.upper()


def _find_avg_col_index(headers: list[str]) -> int | None:
    """Find the index of Average/Close column in value headers."""
    for i, h in enumerate(headers):
        if h and h.strip().lower() in ('average', 'close'):
            return i
    return None


# ── CFM Loader ─────────────────────────────────────────────────────

def load_cfm(con: duckdb.DuckDBPyConnection, dry_run: bool = False) -> tuple[int, int]:
    """Load all CFM files. Returns (n_indicators, n_rows)."""
    cfm_dir = DATA_DIR / 'CFM'
    total_ind = 0
    total_rows = 0

    for fpath in sorted(cfm_dir.glob('*.xlsx')):
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        fname = fpath.stem  # e.g., "DDR", "Channel SSD"

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue

            header = list(all_rows[0])
            val_headers = header[4:]  # after Base Date, Release Date, Time, TZ

            # Determine which column to load
            if fname == 'Index':
                # CFM Index: columns are "DRAM Index", "NAND Index" — load each
                for ci, col_name in enumerate(val_headers):
                    if col_name is None:
                        continue
                    indicator_id = f'CFM_{_slugify(col_name)}'
                    n = _load_standard_rows(
                        con, all_rows[1:], 4 + ci, indicator_id,
                        f'CFM {col_name}', fname, sheet_name, fpath, dry_run
                    )
                    if n > 0:
                        total_ind += 1
                        total_rows += n
                continue

            # Standard CFM sheet: find Average or Close column
            avg_idx = _find_avg_col_index(val_headers)
            if avg_idx is None:
                # Fallback: use the last value column
                avg_idx = len([h for h in val_headers if h]) - 1

            col_abs = 4 + avg_idx  # absolute column index

            # Build indicator_id from file + sheet name
            if fname in sheet_name:
                indicator_id = f'CFM_{_slugify(sheet_name)}'
            else:
                indicator_id = f'CFM_{_slugify(fname)}_{_slugify(sheet_name)}'

            display_name = f'CFM {sheet_name} (Avg)'

            n = _load_standard_rows(
                con, all_rows[1:], col_abs, indicator_id,
                display_name, fname, sheet_name, fpath, dry_run
            )
            if n > 0:
                total_ind += 1
                total_rows += n

        wb.close()

    return total_ind, total_rows


# ── TrendForce Loader ──────────────────────────────────────────────

def load_trendforce(con: duckdb.DuckDBPyConnection, dry_run: bool = False) -> tuple[int, int]:
    """Load all TrendForce files. Returns (n_indicators, n_rows)."""
    tf_dir = DATA_DIR / 'TrendForce'
    total_ind = 0
    total_rows = 0

    for fpath in sorted(tf_dir.glob('*.xlsx')):
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        fname = fpath.stem

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue

            row0 = list(all_rows[0])
            row1 = list(all_rows[1]) if len(all_rows) > 1 else []

            # Detect if this is a two-row header
            is_two_row = (
                len(row1) > 4
                and row1[0] is None  # no date in row 1
                and any(
                    str(v).strip().lower() in ('high', 'low', 'average', 'last avg')
                    for v in row1[4:] if v
                )
            )

            if is_two_row:
                ni, nr = _load_trendforce_multirow(
                    con, all_rows, fname, sheet_name, fpath, dry_run
                )
            else:
                # Single-row header (e.g., DXI index, Li-Ion Battery)
                ni, nr = _load_trendforce_simple(
                    con, all_rows, fname, sheet_name, fpath, dry_run
                )

            total_ind += ni
            total_rows += nr

        wb.close()

    return total_ind, total_rows


def _load_trendforce_multirow(
    con, all_rows, fname, sheet_name, fpath, dry_run
) -> tuple[int, int]:
    """Parse two-row header TrendForce sheets. Extract Average for each product."""
    row0 = list(all_rows[0])  # product names
    row1 = list(all_rows[1])  # High/Low/Average labels
    data_rows = all_rows[2:]

    # Discover product groups from row0[4:]
    products = {}  # {product_name: avg_col_absolute_index}
    val_start = 4
    i = val_start
    while i < len(row0):
        product_name = row0[i]
        if product_name is None:
            i += 1
            continue

        product_name = str(product_name).strip().replace('\xa0', '')

        # Find span: consecutive columns with same product name
        j = i + 1
        while j < len(row0) and row0[j] == row0[i]:
            j += 1

        # Within this span, find the "Average" column from row1
        avg_col = None
        for k in range(i, j):
            if k < len(row1) and row1[k] and str(row1[k]).strip().lower() == 'average':
                avg_col = k
                break

        if avg_col is not None and product_name not in products:
            products[product_name] = avg_col

        i = j

    # Load each product
    n_ind = 0
    n_rows = 0
    for product_name, avg_col in products.items():
        slug = _slugify(product_name)
        sheet_slug = _slugify(sheet_name)

        # Avoid redundancy in naming
        if sheet_slug in slug:
            indicator_id = f'TF_{slug}'
        else:
            indicator_id = f'TF_{sheet_slug}_{slug}'

        # Truncate very long ids
        if len(indicator_id) > 60:
            indicator_id = indicator_id[:60]

        display_name = f'TF {sheet_name} {product_name} (Avg)'

        n = _load_standard_rows(
            con, data_rows, avg_col, indicator_id,
            display_name, fname, sheet_name, fpath, dry_run
        )
        if n > 0:
            n_ind += 1
            n_rows += n

    return n_ind, n_rows


def _load_trendforce_simple(
    con, all_rows, fname, sheet_name, fpath, dry_run
) -> tuple[int, int]:
    """Load single-row header TrendForce sheets (DXI, Li-Ion Battery, etc.)."""
    header = list(all_rows[0])
    val_headers = header[4:]
    data_rows = all_rows[1:]

    n_ind = 0
    n_rows = 0
    for ci, col_name in enumerate(val_headers):
        if col_name is None:
            continue
        col_name_str = str(col_name).strip().replace('\xa0', '')
        if not col_name_str:
            continue

        slug = _slugify(col_name_str)
        sheet_slug = _slugify(sheet_name)

        if sheet_slug in slug:
            indicator_id = f'TF_{slug}'
        else:
            indicator_id = f'TF_{sheet_slug}_{slug}'

        if len(indicator_id) > 60:
            indicator_id = indicator_id[:60]

        display_name = f'TF {sheet_name} {col_name_str}'

        n = _load_standard_rows(
            con, data_rows, 4 + ci, indicator_id,
            display_name, fname, sheet_name, fpath, dry_run
        )
        if n > 0:
            n_ind += 1
            n_rows += n

    return n_ind, n_rows


# ── Shared row loader ──────────────────────────────────────────────

def _load_standard_rows(
    con, data_rows, col_abs, indicator_id, display_name,
    fname, sheet_name, fpath, dry_run
) -> int:
    """Extract and load rows from a standard data block."""
    rows = []
    for row in data_rows:
        if col_abs >= len(row):
            continue

        obs_date = _parse_date(row[0])
        release_date = _parse_date(row[1])
        if obs_date is None:
            if release_date is not None:
                obs_date = release_date
                release_date = None
            else:
                continue

        value_raw = row[col_abs]
        if value_raw is None:
            continue
        try:
            value = float(value_raw)
        except (ValueError, TypeError):
            continue

        rows.append((obs_date, release_date, value))

    if not rows:
        return 0

    print(f'  {indicator_id:55s} {len(rows):>5d} rows', end='')

    if dry_run:
        print(' (dry-run)')
        return len(rows)

    # Upsert dim_indicator
    source = 'CFM' if 'CFM' in str(fpath) else 'TrendForce'
    frequency = 'weekly' if source == 'CFM' else _guess_frequency(rows)
    source_file = str(fpath.relative_to(fpath.parent.parent.parent))

    con.execute('''
        INSERT INTO dim_indicator (
            indicator_id, name, category, subcategory, country,
            frequency, unit, source, collection_method
        ) VALUES (?, ?, 'industry', 'semiconductor', 'GLOBAL', ?, 'usd', ?, 'manual_excel')
        ON CONFLICT (indicator_id) DO UPDATE SET name = EXCLUDED.name
    ''', [indicator_id, display_name, frequency, source])

    # Insert fact rows
    params = [
        (indicator_id, obs, rel, val, 0, source_file)
        for obs, rel, val in rows
    ]
    con.executemany('''
        INSERT INTO fact_indicator_value
            (indicator_id, observation_date, release_date, value, revision, source_file)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT (indicator_id, observation_date, revision) DO UPDATE SET
            value = EXCLUDED.value,
            release_date = EXCLUDED.release_date,
            source_file = EXCLUDED.source_file
    ''', params)

    print()
    return len(rows)


def _guess_frequency(rows: list[tuple]) -> str:
    """Guess frequency from date gaps."""
    if len(rows) < 3:
        return 'irregular'
    dates = sorted(set(r[0] for r in rows))
    if len(dates) < 3:
        return 'irregular'
    gaps = [(dates[i+1] - dates[i]).days for i in range(min(10, len(dates)-1))]
    median_gap = sorted(gaps)[len(gaps)//2]
    if median_gap <= 2:
        return 'daily'
    if median_gap <= 8:
        return 'weekly'
    if median_gap <= 35:
        return 'monthly'
    return 'quarterly'


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description='Load CFM/TrendForce industry data')
    parser.add_argument('--db', default=str(DB_PATH))
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--cfm-only', action='store_true')
    parser.add_argument('--tf-only', action='store_true')
    args = parser.parse_args()

    con = duckdb.connect(args.db) if not args.dry_run else None

    grand_ind = 0
    grand_rows = 0

    if not args.tf_only:
        print('=== CFM ===')
        ni, nr = load_cfm(con, args.dry_run)
        print(f'CFM: {ni} indicators, {nr:,} rows\n')
        grand_ind += ni
        grand_rows += nr

    if not args.cfm_only:
        print('=== TrendForce ===')
        ni, nr = load_trendforce(con, args.dry_run)
        print(f'TrendForce: {ni} indicators, {nr:,} rows\n')
        grand_ind += ni
        grand_rows += nr

    if con:
        con.close()

    print(f'Total: {grand_ind} indicators, {grand_rows:,} rows')


if __name__ == '__main__':
    main()
